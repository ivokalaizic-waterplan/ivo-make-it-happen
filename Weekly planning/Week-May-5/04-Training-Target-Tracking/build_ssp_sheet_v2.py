#!/usr/bin/env python3
"""
SSP Task Sheet v2 — Simplificado
==================================
Sheet limpio para que los SSPs completen:
- Company + SSP
- Targets water/carbon públicos + URL sustainability report
- TT incluido en scope?
- Potencial
- Fecha demo

Pre-poblado con data de plataforma (variables, años, proyectos, targets).
"""

import csv
import json
import sys
import time
import requests
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# --- Config ---
BASE_DIR = Path(__file__).parent
CSV_PATH = Path("/Users/ivokalaizicwp/Documents/waterplan/growth tool/active-licenses-export-2026-05-08.csv")
API_BASE = "https://api.waterplan.com"
TOKEN = sys.argv[1] if len(sys.argv) > 1 else ""
HEADERS = {"Authorization": f"Bearer {TOKEN}"}

SKIP_COMPANIES_API = {
    "demo-waterplan", "demo-waterplan-2", "demo-sales",
    "sandbox-emanuelcasco", "nico-consu-mili-company", "waterplan-og"
}

COMPANY_NAME_TO_API_ID = {
    "Amazon": "amazon", "Amazon Web Services EMEA": "AWS",
    "Coca-Cola | Hellenic": "coca-cola-hellenic", "P&G": "pg2",
    "AB InBev": "abinbev", "EdgeConneX": "edgeconnex", "Alcoa": "alcoa",
    "Coca-Cola | Europacific Partners (CCEP)": "coca-cola-euro-pacific",
    "Colgate-Palmolive": "colgate2", "Metalsa": "metalsa",
    "Coca-Cola | Femsa": "coca-cola-femsa",
    "Coca-Cola | Eurasia & Middle East OU": "coca-cola-euroasia",
    "Constellation Brands": "constellation-brands", "Bocar": "bocargroup",
    "Coca-Cola | INSWA OU": "coca-cola-inswa", "Premier Foods": "premier-foods",
    "Sigma": "sigma", "Coca-Cola | Africa OU": "coca-cola-africa-ou",
    "AB InBev | Ambev": "ambev",
    "Coca-Cola | Arca Continental": "coca-cola-arca-continental",
    "Remy Cointreau": "remy-cointreau", "Pluspetrol": "pluspetrol",
    "Conagra Brands": "conagra-brands", "Coca Cola System": "tccc-apac",
    "Citrofrut": "citrofrut", "Manulife": "manulife",
    "Porto do Açu": "porto-do-acu", "Baiada": "baiada",
    "Driscoll's": "driscolls", "Dole Packaged Foods": "dole",
    "Gestamp": "gestamp", "Heineken": "heineken", "Inditex": "inditex",
    "COSTCO": "costcomx",
    "Coca-Cola | Equatorial Coca-Cola Bottling Co. (ECCBC)": "coca-cola-equatorial",
    "PepsiCo": "pepsico", "PepsiCo | Postobón": "pepsico",
    "Coca-Cola | Andina": "coca-cola-andina", "Estee Lauder": "estee-lauder",
    "AB InBev | Quilmes": "ambev", "Femsa Corporate": "femsa-corporate",
    "Brown Forman": "herradura-brown-forman", "Danone": "danone",
    "Grupo Boticário": "grupo-boticario", "Huhtamaki": "huhtamaki",
    "Natura": "natura-v2", "Nestle": "nestle", "Omya": "omya",
    "Orica": "orica", "PAE": "pae", "Pan American Energy": "pae",
    "Petrobras": "petrobras", "Rotoplas": "rotoplas",
    "Suntory": "suntory", "Tronox": "tronox", "Unilever": "unilever",
}


def api_get(path, params=None):
    try:
        resp = requests.get(f"{API_BASE}{path}", headers=HEADERS, params=params, timeout=30)
        if resp.status_code == 200:
            return resp.json()
        return None
    except:
        return None


def get_api_id(company_name):
    if company_name in COMPANY_NAME_TO_API_ID:
        return COMPANY_NAME_TO_API_ID[company_name]
    return None


def audit_company(cid):
    """Full audit of a single company."""
    # Catalog
    catalog = api_get(f"/v3/data-gathering/{cid}/catalog", {"includeAvailability": "true"})
    if not isinstance(catalog, list):
        catalog = []

    vars_with_data = []
    min_year = max_year = None
    categories = set()
    for var in catalog:
        avail = var.get("availability")
        if avail and avail.get("from"):
            vars_with_data.append(var.get("title", var.get("variableId", "?")))
            cat = var.get("category")
            if cat:
                categories.add(cat)
            fy = int(avail["from"][:4])
            ty = int(avail.get("to", avail["from"])[:4])
            if min_year is None or fy < min_year: min_year = fy
            if max_year is None or ty > max_year: max_year = ty

    # Projects (ITF) - handles both dict and list responses
    projects_raw = api_get(f"/v3/companies/{cid}/inside-the-fence/projects")
    if isinstance(projects_raw, dict):
        projects = projects_raw.get("projects", [])
    elif isinstance(projects_raw, list):
        projects = projects_raw
    else:
        projects = []

    # Count by phase
    phases = {}
    for p in projects:
        ph = p.get("phase", "unknown")
        phases[ph] = phases.get(ph, 0) + 1

    # Targets
    targets_raw = api_get(f"/v3/target-tracking/v2/{cid}/targets", {"limit": 100})
    targets = []
    if isinstance(targets_raw, dict):
        targets = targets_raw.get("targets", [])
    elif isinstance(targets_raw, list):
        targets = targets_raw

    target_details = []
    for t in targets:
        name = t.get("name", t.get("metricId", "?"))
        types = t.get("types", [])
        baseline = t.get("baseline", {})
        target_info = t.get("target", {})
        b_yr = baseline.get("date", {}).get("year", "?") if isinstance(baseline.get("date"), dict) else "?"
        t_yr = target_info.get("date", {}).get("year", "?") if isinstance(target_info.get("date"), dict) else "?"
        target_details.append(f"{name} ({b_yr}->{t_yr})")

    year_range = f"{min_year}-{max_year}" if min_year else "Sin data"

    return {
        "has_data": len(vars_with_data) > 0,
        "vars_count": len(vars_with_data),
        "categories": ", ".join(sorted(categories)),
        "year_range": year_range,
        "projects_total": len(projects),
        "projects_phases": phases,
        "targets_count": len(targets),
        "target_details": "; ".join(target_details),
    }


# --- Load licenses ---
print("Loading licenses...")
licenses = []
with open(CSV_PATH, "r", encoding="utf-8") as f:
    for row in csv.DictReader(f):
        licenses.append(row)

# Group by company
company_groups = {}
for lic in licenses:
    cname = lic["Company"].strip()
    if cname not in company_groups:
        company_groups[cname] = {"company": cname, "ssps": set(), "licenses": []}
    g = company_groups[cname]
    ssp = lic.get("SSP", "").strip()
    if ssp:
        g["ssps"].add(ssp)
    g["licenses"].append(lic)

# --- Audit all companies ---
print(f"\nAuditing {len(company_groups)} companies from CRM...")
audit_results = {}

for i, (cname, g) in enumerate(sorted(company_groups.items())):
    api_id = get_api_id(cname)
    if api_id and api_id not in SKIP_COMPANIES_API:
        print(f"  [{i+1}/{len(company_groups)}] {cname} ({api_id})...", end=" ", flush=True)
        result = audit_company(api_id)
        audit_results[cname] = result
        icon = "Y" if result["has_data"] else "N"
        print(f"data={icon} vars={result['vars_count']} projects={result['projects_total']} targets={result['targets_count']}")
        time.sleep(0.2)
    else:
        print(f"  [{i+1}/{len(company_groups)}] {cname} — sin instancia API")
        audit_results[cname] = None


# --- Build Excel ---
print("\nBuilding Excel...")
wb = openpyxl.Workbook()

# Styles
header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True, size=10)
auto_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
ssp_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
wrap = Alignment(wrap_text=True, vertical="top")

ws = wb.active
ws.title = "SSP Task - Target Tracking"

# --- Section headers (row 1) ---
sections = [
    (1, 2, "DATOS CUENTA", header_fill),
    (3, 8, "DATA EN PLATAFORMA (auto - no tocar)", auto_fill),
    (9, 16, "SSP COMPLETA ESTO", ssp_fill),
]
for sc, ec, title, fill in sections:
    ws.merge_cells(start_row=1, start_column=sc, end_row=1, end_column=ec)
    cell = ws.cell(row=1, column=sc, value=title)
    cell.fill = fill
    cell.font = Font(bold=True, size=11, color="FFFFFF" if fill == header_fill else "000000")
    cell.alignment = Alignment(horizontal="center")
    cell.border = thin_border
    for c in range(sc + 1, ec + 1):
        ws.cell(row=1, column=c).fill = fill
        ws.cell(row=1, column=c).border = thin_border

# --- Column headers (row 2) ---
columns = [
    # Account data
    ("Company", 25),                              # A
    ("SSP", 22),                                   # B
    # Platform data (auto)
    ("Data cargada?", 12),                         # C
    ("Variables con data", 10),                    # D
    ("Rango de años", 12),                         # E
    ("Proyectos cargados", 10),                    # F
    ("Fases proyectos", 25),                       # G
    ("Targets en plataforma", 30),                 # H
    # SSP fills
    ("TT incluido en scope? (SI/NO/NO SE)", 16),   # I
    ("Water Target 1", 35),                        # J
    ("Water Target 2", 35),                        # K
    ("Carbon Target 1", 35),                       # L
    ("Carbon Target 2", 35),                       # M
    ("URL Sustainability Report", 35),             # N
    ("Ve utilidad en TT? (SI/NO + por qué)", 30),  # O
    ("Fecha propuesta demo (con Ivo)", 18),        # P
]

for col, (name, width) in enumerate(columns, 1):
    cell = ws.cell(row=2, column=col, value=name)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", wrap_text=True)
    cell.border = thin_border
    ws.column_dimensions[get_column_letter(col)].width = width

# --- Data rows ---
# Sort: companies with data first (by projects desc, then vars desc), then without
def row_sort(item):
    cname, g = item
    a = audit_results.get(cname)
    if a is None:
        return (2, 0, 0)  # no API instance → last
    if not a["has_data"]:
        return (1, 0, 0)  # no data → middle
    return (0, -a["projects_total"], -a["vars_count"])  # has data → first, by projects

row = 3
for cname, g in sorted(company_groups.items(), key=row_sort):
    a = audit_results.get(cname)
    ssps = ", ".join(sorted(g["ssps"]))

    # A-B: Account
    ws.cell(row=row, column=1, value=cname).border = thin_border
    ws.cell(row=row, column=2, value=ssps).border = thin_border

    # C-H: Platform data
    if a is not None:
        # C: Data cargada?
        c = ws.cell(row=row, column=3, value="SI" if a["has_data"] else "NO")
        c.fill = green_fill if a["has_data"] else red_fill
        c.font = Font(bold=True)
        c.border = thin_border

        # D: Variables count
        ws.cell(row=row, column=4, value=a["vars_count"]).border = thin_border

        # E: Year range
        ws.cell(row=row, column=5, value=a["year_range"]).border = thin_border

        # F: Projects count
        proj_cell = ws.cell(row=row, column=6, value=a["projects_total"])
        proj_cell.border = thin_border
        if a["projects_total"] > 0:
            proj_cell.fill = green_fill

        # G: Project phases
        if a["projects_phases"]:
            phase_str = ", ".join(f"{k}: {v}" for k, v in sorted(a["projects_phases"].items()))
        else:
            phase_str = ""
        ws.cell(row=row, column=7, value=phase_str).border = thin_border

        # H: Targets detail
        ws.cell(row=row, column=8, value=a["target_details"]).border = thin_border
    else:
        c = ws.cell(row=row, column=3, value="SIN INSTANCIA")
        c.fill = gray_fill
        c.border = thin_border
        for col in range(4, 9):
            ws.cell(row=row, column=col, value="").border = thin_border

    # I-P: SSP fills (empty)
    for col in range(9, 17):
        cell = ws.cell(row=row, column=col, value="")
        cell.border = thin_border
        cell.alignment = wrap

    row += 1

# Data validations
dv_tt = DataValidation(type="list", formula1='"SI,NO,NO SE"', allow_blank=True)
dv_tt.error = "Seleccionar SI, NO, o NO SE"
ws.add_data_validation(dv_tt)
dv_tt.add(f"I3:I{row-1}")

dv_util = DataValidation(type="list", formula1='"SI,NO"', allow_blank=True)
ws.add_data_validation(dv_util)
dv_util.add(f"O3:O{row-1}")

ws.freeze_panes = "A3"

# --- Sheet 2: Instrucciones ---
ws2 = wb.create_sheet("Instrucciones")
instr = [
    ("TARGET TRACKING - TAREA PARA SSPs", True, 14),
    ("", False, 10),
    ("Fecha: 2026-05-08 | Owner: Ivo Kalaizic | Deadline: 2026-05-23", False, 11),
    ("", False, 10),
    ("OBJETIVO:", True, 12),
    ("Completar las columnas amarillas (I-P) para TODOS tus clientes.", False, 10),
    ("", False, 10),
    ("COLUMNAS A COMPLETAR:", True, 12),
    ("", False, 10),
    ("I - TT INCLUIDO EN SCOPE? (OBLIGATORIO)", True, 11),
    ("   Validar si Target Tracking está incluido en el contrato.", False, 10),
    ("   Seleccionar: SI / NO / NO SE", False, 10),
    ("   Si no sabés, consultá con tu AE y después actualizá.", False, 10),
    ("", False, 10),
    ("J-K - WATER TARGETS PÚBLICOS", True, 11),
    ("   Buscar en el sustainability report los compromisos de AGUA.", False, 10),
    ("   Escribir el target completo. Ej:", False, 10),
    ("   'Reducir water intensity 20% para 2030 vs baseline 2020'", False, 10),
    ("   'Reducir total withdrawals 15% para 2025'", False, 10),
    ("   'Water use ratio de 1.2 L/L para 2030'", False, 10),
    ("", False, 10),
    ("L-M - CARBON TARGETS PÚBLICOS", True, 11),
    ("   Buscar en el sustainability report los compromisos de CARBONO/ENERGÍA.", False, 10),
    ("   Ej: 'Scope 1+2 reducir 42% para 2030 (SBTi)'", False, 10),
    ("   'Net zero para 2050'", False, 10),
    ("", False, 10),
    ("N - URL SUSTAINABILITY REPORT (OBLIGATORIO si ponen targets)", True, 11),
    ("   Link directo al documento fuente.", False, 10),
    ("", False, 10),
    ("O - VE UTILIDAD EN TT?", True, 11),
    ("   SI o NO + justificación breve de por qué sí o por qué no.", False, 10),
    ("   Ej: 'SI - el cliente tiene reuniones trimestrales de progreso'", False, 10),
    ("   Ej: 'NO - el cliente es solo Water Risk, no mide targets'", False, 10),
    ("", False, 10),
    ("P - FECHA PROPUESTA DEMO", True, 11),
    ("   Si TT está en scope Y hay data cargada:", False, 10),
    ("   Proponer fecha para demo con el cliente.", False, 10),
    ("   IVO KALAIZIC DEBE SER INVITADO A LA REUNIÓN.", False, 10),
    ("", False, 10),
    ("COLUMNAS PRE-POBLADAS (NO TOCAR):", True, 12),
    ("   C: Si el cliente tiene data en Data Gathering", False, 10),
    ("   D: Cantidad de variables con data", False, 10),
    ("   E: Rango de años cargados (ej: 2020-2025)", False, 10),
    ("   F: Cantidad de proyectos cargados en la plataforma", False, 10),
    ("   G: Distribución de fases de los proyectos", False, 10),
    ("   H: Targets que ya existen en la plataforma (nombre + periodo)", False, 10),
]
for r, (text, bold, size) in enumerate(instr, 1):
    cell = ws2.cell(row=r, column=1, value=text)
    cell.font = Font(bold=bold, size=size)
    if bold and size >= 12:
        cell.fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
ws2.column_dimensions["A"].width = 80

# --- Save ---
out = BASE_DIR / "outputs" / f"SSP_Task_TT_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
out.parent.mkdir(exist_ok=True)
wb.save(str(out))

# Summary
with_data = sum(1 for a in audit_results.values() if a and a["has_data"])
with_projects = sum(1 for a in audit_results.values() if a and a["projects_total"] > 0)
with_targets = sum(1 for a in audit_results.values() if a and a["targets_count"] > 0)
no_instance = sum(1 for a in audit_results.values() if a is None)

print(f"\n{'='*60}")
print(f"RESUMEN")
print(f"{'='*60}")
print(f"Total companies: {len(company_groups)}")
print(f"Con data en plataforma: {with_data}")
print(f"Con proyectos cargados: {with_projects}")
print(f"Con targets creados: {with_targets}")
print(f"Sin instancia API: {no_instance}")
print(f"\nExcel: {out}")
