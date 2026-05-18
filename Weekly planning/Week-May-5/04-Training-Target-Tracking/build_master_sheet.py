#!/usr/bin/env python3
"""
Master Sheet Builder: TT Readiness x Active Licenses
=====================================================
Cruza:
1. CSV de licencias activas (SSP, AE, package, ARR)
2. Auditoría de Data Gathering (variables, años, targets, proyectos)

Genera un Excel maestro para que los SSPs completen.
"""

import csv
import json
import sys
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# --- Config ---
BASE_DIR = Path(__file__).parent
CSV_PATH = Path("/Users/ivokalaizicwp/Documents/waterplan/growth tool/active-licenses-export-2026-05-08.csv")
AUDIT_JSON = sorted(BASE_DIR.glob("outputs/TT_Readiness_Raw_*.json"))[-1]  # Latest audit

# --- Load data ---
print(f"Loading licenses from: {CSV_PATH}")
print(f"Loading audit from: {AUDIT_JSON}")

# Load licenses
licenses = []
with open(CSV_PATH, "r", encoding="utf-8") as f:
    reader = csv.DictReader(f)
    for row in reader:
        licenses.append(row)

# Load audit
with open(AUDIT_JSON, "r") as f:
    audit_data = json.load(f)

# Build audit lookup by company_id
audit_lookup = {r["company_id"]: r for r in audit_data}

# --- Map license companies to API company IDs ---
# This mapping handles the mismatch between CRM names and API IDs
COMPANY_NAME_TO_API_ID = {
    "Amazon": "amazon",
    "Amazon Web Services EMEA": "AWS",
    "Coca-Cola | Hellenic": "coca-cola-hellenic",
    "P&G": "pg2",
    "AB InBev": "abinbev",
    "EdgeConneX": "edgeconnex",
    "Alcoa": "alcoa",
    "Coca-Cola | Europacific Partners (CCEP)": "coca-cola-euro-pacific",
    "Colgate-Palmolive": "colgate2",
    "Metalsa": "metalsa",
    "Coca-Cola | Femsa": "coca-cola-femsa",
    "Coca-Cola | Eurasia & Middle East OU": "coca-cola-euroasia",
    "Constellation Brands": "constellation-brands",
    "Bocar": "bocargroup",
    "Coca-Cola | INSWA OU": "coca-cola-inswa",
    "Premier Foods": "premier-foods",
    "Sigma": "sigma",
    "Coca-Cola | Africa OU": "coca-cola-africa-ou",
    "AB InBev | Ambev": "ambev",
    "Coca-Cola | Arca Continental": "coca-cola-arca-continental",
    "Remy Cointreau": "remy-cointreau",
    "Pluspetrol": "pluspetrol",
    "Conagra Brands": "conagra-brands",
    "Coca Cola System": "tccc-apac",
    "Citrofrut": "citrofrut",
    "Manulife": "manulife",
    "Porto do Açu": "porto-do-acu",
    "Baiada": "baiada",
    "Driscoll's": "driscolls",
    "Dole Packaged Foods": "dole",
    "Gestamp": "gestamp",
    "Heineken": "heineken",
    "Inditex": "inditex",
    "COSTCO": "costcomx",
    "Coca-Cola | Equatorial Coca-Cola Bottling Co. (ECCBC)": "coca-cola-equatorial",
    "PepsiCo": "pepsico",
    "PepsiCo | Postobón": "pepsico",
    "PepsiCo | CBC": None,
    "PepsiCo | CCU": None,
    "Coca-Cola | Andina": "coca-cola-andina",
    "Estee Lauder": "estee-lauder",
    "AB InBev | Quilmes": "ambev",
    "Coca-Cola | Innocent Drinks": None,
    "Coca-Cola | SFBT": None,
    "Coca-Cola | Delta corporations": None,
    "Coca-Cola | CCI": None,
    "Coca-Cola | Phoenix Bev": None,
    "Coca-Cola | Refriango": None,
    "Femsa Corporate": "femsa-corporate",
    # New companies with access
    "Brown Forman": "herradura-brown-forman",
    "Danone": "danone",
    "Grupo Boticário": "grupo-boticario",
    "Huhtamaki": "huhtamaki",
    "Natura": "natura-v2",
    "Nestle": "nestle",
    "Omya": "omya",
    "Orica": "orica",
    "PAE": "pae",
    "Petrobras": "petrobras",
    "Rotoplas": "rotoplas",
    "Suntory": "suntory",
    "Tronox": "tronox",
    "Unilever": "unilever",
    "HENKEL": None,
    "Telecom": None,
    "San Miguel Global": None,
    "Anadolu Efes": None,
    "Imerys": None,
    "AG Barr": None,
    "Accenture": None,
    "Equinix": None,
    "Froneri": None,
    "Baxter": None,
    "The Compleat": None,
    "Toppan": None,
    "Syngenta": None,
    "BAT": None,
    "Pan American Energy": "pae",
}


def get_api_id(company_name):
    """Try to match a license company name to an API company ID."""
    if company_name in COMPANY_NAME_TO_API_ID:
        return COMPANY_NAME_TO_API_ID[company_name]
    # Fuzzy fallback: try lowercase slug
    slug = company_name.lower().replace(" ", "-").replace("|", "").replace("  ", "-").strip("-")
    for aid in audit_lookup:
        if aid in slug or slug in aid:
            return aid
    return None


# --- Aggregate licenses by unique company (for SSP assignment) ---
# Group by company name, collecting all licenses
company_groups = {}
for lic in licenses:
    cname = lic["Company"].strip()
    if cname not in company_groups:
        company_groups[cname] = {
            "company": cname,
            "licenses": [],
            "ssps": set(),
            "aes": set(),
            "packages": set(),
            "has_tt_in_scope": False,
            "tt_license_names": [],
            "total_arr": 0,
        }
    g = company_groups[cname]
    g["licenses"].append(lic)
    ssp = lic.get("SSP", "").strip()
    ae = lic.get("AE", "").strip()
    if ssp:
        g["ssps"].add(ssp)
    if ae:
        g["aes"].add(ae)

    pkg = lic.get("License Package", "")
    for p in pkg.split(";"):
        p = p.strip()
        if p:
            g["packages"].add(p)
    if "Target Tracking" in pkg or "Carbon" in pkg:
        g["has_tt_in_scope"] = True
        g["tt_license_names"].append(lic["Active License Name"].strip())

    try:
        arr = float(lic.get("Current ARR", "0").replace(",", ""))
    except ValueError:
        arr = 0
    g["total_arr"] += arr


# --- Build Excel ---
wb = openpyxl.Workbook()

# Styles
header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True, size=10)
section_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
section_font = Font(color="FFFFFF", bold=True, size=10)
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
light_blue_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
light_gray_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
wrap_align = Alignment(wrap_text=True, vertical="top")

# ============================================================
# SHEET 1: Master View
# ============================================================
ws = wb.active
ws.title = "TT Activation Master"

# Column groups
# A: Pre-populated (we fill)
# B: SSP fills
headers = [
    # --- Pre-populated section ---
    ("Company", 22),                        # A
    ("SSP Owner(s)", 20),                   # B
    ("AE", 18),                             # C
    ("Total ARR", 12),                      # D
    ("License Packages (ref)", 25),         # E
    # --- Platform Data (auto from API) ---
    ("Data en Plataforma?", 14),            # F
    ("Variables con Data", 12),             # G
    ("Categorias DG", 20),                  # H
    ("Rango de Años", 14),                  # I
    ("Targets en Plataforma", 12),          # J
    ("Targets Detalle", 30),                # K
    ("Proyectos ITF", 12),                  # L
    ("Readiness Score", 12),                # M
    ("Readiness Status", 14),               # N
    # --- SSP debe completar ---
    ("TT incluido en scope del contrato? (SI/NO/NO SE)", 20),  # O
    ("Water Target 1 (público)", 30),       # P
    ("Water Target 2 (público)", 30),       # Q
    ("Water Target 3 (público)", 30),       # R
    ("Carbon Target 1 (público)", 30),      # S
    ("Carbon Target 2 (público)", 30),      # T
    ("URL Sustainability Report", 35),      # U
    ("Ve potencial TT? (1-5)", 12),         # V
    ("Fecha propuesta demo", 16),           # W
    ("Notas / Blockers", 30),               # X
]

# Section headers row
section_headers = [
    (1, 5, "DATOS DE LICENCIA (pre-poblado - no tocar)", section_fill),
    (6, 14, "DATA EN PLATAFORMA (auto - API - no tocar)", light_blue_fill),
    (15, 24, "SSP DEBE COMPLETAR", yellow_fill),
]

# Write section headers (row 1)
for start_col, end_col, title, fill in section_headers:
    ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
    cell = ws.cell(row=1, column=start_col, value=title)
    cell.fill = fill
    cell.font = Font(bold=True, size=11, color="FFFFFF" if fill != yellow_fill else "000000")
    cell.alignment = Alignment(horizontal="center")
    cell.border = thin_border
    for c in range(start_col + 1, end_col + 1):
        ws.cell(row=1, column=c).fill = fill
        ws.cell(row=1, column=c).border = thin_border

# Write column headers (row 2)
for col, (header, width) in enumerate(headers, 1):
    cell = ws.cell(row=2, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", wrap_text=True)
    cell.border = thin_border
    ws.column_dimensions[get_column_letter(col)].width = width

# Sort companies: by readiness score desc (from API), then by ARR desc
def sort_key(g):
    api_id = get_api_id(g["company"])
    audit = audit_lookup.get(api_id, {}) if api_id else {}
    score = audit.get("readiness_score", -1)
    return (-score, -g["total_arr"])

sorted_companies = sorted(company_groups.values(), key=sort_key)

# Write data rows
row = 3
no_api_access = []

for g in sorted_companies:
    cname = g["company"]
    api_id = get_api_id(cname)
    audit = audit_lookup.get(api_id, {}) if api_id else {}

    # Pre-populated columns (A-E)
    ws.cell(row=row, column=1, value=cname).border = thin_border
    ws.cell(row=row, column=2, value=", ".join(sorted(g["ssps"]))).border = thin_border
    ws.cell(row=row, column=3, value=", ".join(sorted(g["aes"]))).border = thin_border

    arr_cell = ws.cell(row=row, column=4, value=g["total_arr"])
    arr_cell.number_format = '#,##0'
    arr_cell.border = thin_border

    ws.cell(row=row, column=5, value="; ".join(sorted(g["packages"]))).border = thin_border

    # Platform data columns (F-N)
    if audit:
        has_data = audit.get("variables_with_data", 0) > 0
        data_cell = ws.cell(row=row, column=6, value="SI" if has_data else "NO")
        data_cell.fill = green_fill if has_data else red_fill
        data_cell.font = Font(bold=True)
        data_cell.border = thin_border

        ws.cell(row=row, column=7, value=audit.get("variables_with_data", 0)).border = thin_border
        ws.cell(row=row, column=8, value=audit.get("categories", "")).border = thin_border
        ws.cell(row=row, column=9, value=audit.get("year_range", "N/A")).border = thin_border
        ws.cell(row=row, column=10, value=audit.get("targets_count", 0)).border = thin_border

        target_names = audit.get("target_names", [])
        ws.cell(row=row, column=11, value=", ".join(target_names) if target_names else "").border = thin_border
        ws.cell(row=row, column=12, value=audit.get("projects_count", 0)).border = thin_border
        ws.cell(row=row, column=13, value=audit.get("readiness_score", 0)).border = thin_border

        status = audit.get("readiness_status", "?")
        status_cell = ws.cell(row=row, column=14, value=status)
        status_cell.border = thin_border
        if status == "READY":
            status_cell.fill = green_fill
        elif status == "PARTIAL":
            status_cell.fill = yellow_fill
        else:
            status_cell.fill = red_fill
    else:
        # No API access
        no_access_cell = ws.cell(row=row, column=6, value="SIN INSTANCIA")
        no_access_cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        no_access_cell.border = thin_border
        for c in range(7, 15):
            ws.cell(row=row, column=c, value="").border = thin_border
        if api_id is None:
            no_api_access.append(cname)

    # SSP columns (O-X) - empty, to be filled
    for c in range(15, 25):
        cell = ws.cell(row=row, column=c, value="")
        cell.border = thin_border
        cell.alignment = wrap_align

    row += 1

# Add data validations
dv_yesno = DataValidation(type="list", formula1='"SI,NO,NO SE"', allow_blank=True)
dv_yesno.error = "Seleccionar SI, NO, o NO SE"
ws.add_data_validation(dv_yesno)
dv_yesno.add(f"O3:O{row-1}")  # TT en scope (col O)

dv_score = DataValidation(type="list", formula1='"1,2,3,4,5"', allow_blank=True)
ws.add_data_validation(dv_score)
dv_score.add(f"V3:V{row-1}")  # Potencial 1-5 (col V)

# Freeze panes
ws.freeze_panes = "A3"

# ============================================================
# SHEET 2: By SSP
# ============================================================
ws2 = wb.create_sheet("By SSP")

# Get unique SSPs
all_ssps = set()
for g in company_groups.values():
    all_ssps.update(g["ssps"])
all_ssps = sorted(all_ssps)

ssp_headers = ["SSP", "Total Clients", "Clients with TT in Scope",
               "Clients with Data", "Clients READY", "Client List"]
for col, h in enumerate(ssp_headers, 1):
    cell = ws2.cell(row=1, column=col, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.border = thin_border

for r, ssp in enumerate(all_ssps, 2):
    ssp_companies = [g for g in company_groups.values() if ssp in g["ssps"]]
    tt_scope = [g for g in ssp_companies if g["has_tt_in_scope"]]
    with_data = []
    ready_count = 0
    for g in ssp_companies:
        api_id = get_api_id(g["company"])
        if api_id and api_id in audit_lookup:
            a = audit_lookup[api_id]
            if a.get("variables_with_data", 0) > 0:
                with_data.append(g["company"])
            if a.get("readiness_status") == "READY":
                ready_count += 1

    ws2.cell(row=r, column=1, value=ssp).border = thin_border
    ws2.cell(row=r, column=2, value=len(ssp_companies)).border = thin_border
    ws2.cell(row=r, column=3, value=len(tt_scope)).border = thin_border
    ws2.cell(row=r, column=4, value=len(with_data)).border = thin_border
    ws2.cell(row=r, column=5, value=ready_count).border = thin_border
    ws2.cell(row=r, column=6, value=", ".join(g["company"] for g in ssp_companies)).border = thin_border

ws2.freeze_panes = "A2"
for col in range(1, 7):
    ws2.column_dimensions[get_column_letter(col)].width = [20, 12, 18, 14, 14, 60][col-1]

# ============================================================
# SHEET 3: Instrucciones para SSPs
# ============================================================
ws3 = wb.create_sheet("Instrucciones")

instructions = [
    ("TARGET TRACKING ACTIVATION - INSTRUCCIONES PARA SSPs", True, 14),
    ("", False, 10),
    ("Fecha: 2026-05-08 | Owner: Ivo Kalaizic | Deadline: 2026-05-23", False, 10),
    ("", False, 10),
    ("OBJETIVO:", True, 12),
    ("Cada SSP debe completar las columnas amarillas (O-X) del sheet 'TT Activation Master'", False, 10),
    ("para TODOS sus clientes. Esto nos permite determinar quiénes están listos para", False, 10),
    ("una demo de Target Tracking en plataforma.", False, 10),
    ("", False, 10),
    ("QUÉ YA ESTÁ PRE-POBLADO (NO TOCAR):", True, 12),
    ("- Columnas A-E: Datos de licencia (del CRM)", False, 10),
    ("- Columnas F-N: Data en plataforma (consultada desde la API el 2026-05-08)", False, 10),
    ("  Esto muestra si el cliente tiene data cargada, qué variables, en qué años,", False, 10),
    ("  si ya tiene targets creados en la plataforma, y un score de readiness.", False, 10),
    ("", False, 10),
    ("QUÉ DEBEN COMPLETAR (columnas O-X):", True, 12),
    ("", False, 10),
    ("1. TT INCLUIDO EN SCOPE DEL CONTRATO? (columna O) - OBLIGATORIO:", True, 11),
    ("   Seleccionar SI, NO, o NO SE.", False, 10),
    ("   Validar si Target Tracking está incluido en el scope del contrato del cliente.", False, 10),
    ("   Si no estás seguro, consultar con el AE y luego actualizar.", False, 10),
    ("   IMPORTANTE: La columna E muestra los paquetes de licencia como referencia,", False, 10),
    ("   pero NO confiar ciegamente en eso. Validar por cuenta propia.", False, 10),
    ("", False, 10),
    ("2. WATER TARGETS PÚBLICOS (columnas P-R):", True, 11),
    ("   Buscar en el sustainability report del cliente qué compromisos ambientales de AGUA tiene.", False, 10),
    ("   Escribir el target completo. Ejemplos:", False, 10),
    ("   'Reducir water intensity 20% para 2030 vs baseline 2020'", False, 10),
    ("   'Reducir total withdrawals 15% para 2025'", False, 10),
    ("   'Lograr water use ratio de 1.2 L/L para 2030'", False, 10),
    ("   Si tiene más de 3, agregar en Notas.", False, 10),
    ("", False, 10),
    ("3. CARBON TARGETS PÚBLICOS (columnas S-T):", True, 11),
    ("   Igual que agua pero para compromisos de carbono/energía.", False, 10),
    ("   Ejemplos: 'Scope 1+2 reducir 42% para 2030 (SBTi)'", False, 10),
    ("   'Net zero para 2050'", False, 10),
    ("   'Reducir energy intensity 25% para 2030'", False, 10),
    ("", False, 10),
    ("4. URL SUSTAINABILITY REPORT (columna U):", True, 11),
    ("   Link directo al documento de donde sacaron los targets.", False, 10),
    ("   OBLIGATORIO si completan targets. Esto es la fuente de verdad.", False, 10),
    ("", False, 10),
    ("5. POTENCIAL TT (columna V):", True, 11),
    ("   Del 1 al 5, qué tan probable/interesante es que este cliente use TT.", False, 10),
    ("   5 = el cliente lo necesita urgente, 1 = no veo potencial.", False, 10),
    ("", False, 10),
    ("6. FECHA PROPUESTA DEMO (columna W):", True, 11),
    ("   Si TT está en scope Y hay data en plataforma:", False, 10),
    ("   Proponer fecha para demo de TT con el cliente. IVO DEBE SER INVITADO.", False, 10),
    ("", False, 10),
    ("RECORDATORIOS DEL TRAINING:", True, 12),
    ("- Para que TT funcione necesitamos: Data en DG + Proyectos con site/start date/impact", False, 10),
    ("- El caso base usa todos los proyectos que NO están en draft/cancel", False, 10),
    ("- Se pueden crear hasta 5 escenarios por target", False, 10),
    ("- Managed Tags ayudan a categorizar proyectos (hasta 30 tags)", False, 10),
    ("- Tommy Gómez es owner de data uploads → canal de Slack", False, 10),
    ("", False, 10),
    ("FLUJO ESPERADO:", True, 12),
    ("1. SSP completa este sheet (deadline: 16 mayo)", False, 10),
    ("2. Ivo revisa y prioriza clientes para demo", False, 10),
    ("3. SSP agenda call de validación con Ivo", False, 10),
    ("4. SSP hace demo de TT al cliente (Ivo invitado)", False, 10),
    ("5. SSP completa form de feedback post-demo", False, 10),
]

for r, (text, bold, size) in enumerate(instructions, 1):
    cell = ws3.cell(row=r, column=1, value=text)
    cell.font = Font(bold=bold, size=size)
    if bold and size >= 12:
        cell.fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

ws3.column_dimensions["A"].width = 90

# ============================================================
# SHEET 4: Companies sin acceso API
# ============================================================
ws4 = wb.create_sheet("Sin Acceso API")
ws4.cell(row=1, column=1, value="Company (CRM)").font = Font(bold=True)
ws4.cell(row=1, column=2, value="Tiene instancia en plataforma?").font = Font(bold=True)
ws4.cell(row=1, column=3, value="Notas").font = Font(bold=True)

# Find all companies without API match
no_match = set()
for g in company_groups.values():
    api_id = get_api_id(g["company"])
    if api_id is None:
        no_match.add(g["company"])

for r, cname in enumerate(sorted(no_match), 2):
    ws4.cell(row=r, column=1, value=cname)
    ws4.cell(row=r, column=2, value="VERIFICAR")
    ws4.cell(row=r, column=3, value="Agregar acceso a Ivo si existe")

ws4.column_dimensions["A"].width = 40
ws4.column_dimensions["B"].width = 25
ws4.column_dimensions["C"].width = 35


# --- Save ---
output_path = BASE_DIR / "outputs" / f"TT_Master_Sheet_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
output_path.parent.mkdir(exist_ok=True)
wb.save(str(output_path))
print(f"\n✅ Master sheet saved to: {output_path}")

# --- Summary ---
print(f"\n📊 RESUMEN:")
print(f"   Total companies (CRM): {len(company_groups)}")
print(f"   Con TT en scope de contrato: {sum(1 for g in company_groups.values() if g['has_tt_in_scope'])}")
print(f"   Con data en plataforma: {sum(1 for g in company_groups.values() if get_api_id(g['company']) and audit_lookup.get(get_api_id(g['company']), {}).get('variables_with_data', 0) > 0)}")
print(f"   Sin match en API: {len(no_match)}")
if no_match:
    print(f"   → Verificar acceso para: {', '.join(sorted(no_match))}")
