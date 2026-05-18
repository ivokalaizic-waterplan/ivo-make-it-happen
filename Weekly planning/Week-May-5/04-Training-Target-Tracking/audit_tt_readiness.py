#!/usr/bin/env python3
"""
Target Tracking Readiness Audit
================================
Recorre todas las companies de Waterplan y genera un reporte de:
- Data cargada en Data Gathering (variables, años)
- Proyectos ITF cargados (cantidad, fases, métricas de impacto)
- Targets existentes en Target Tracking
- Readiness score para habilitar Target Tracking

Output: Excel con el reporte completo + resumen en consola
"""

import json
import sys
import time
import requests
from datetime import datetime
from pathlib import Path

API_BASE = "https://api.waterplan.com"
TOKEN = sys.argv[1] if len(sys.argv) > 1 else ""

HEADERS = {"Authorization": f"Bearer {TOKEN}"}

# Companies to skip (internal/demo/sandbox)
SKIP_COMPANIES = {
    "demo-waterplan", "demo-waterplan-2", "demo-sales",
    "sandbox-emanuelcasco", "nico-consu-mili-company",
    "waterplan-og"
}


def api_get(path, params=None):
    """GET request with error handling and rate limiting."""
    url = f"{API_BASE}{path}"
    try:
        resp = requests.get(url, headers=HEADERS, params=params, timeout=30)
        if resp.status_code == 200:
            return resp.json()
        elif resp.status_code == 403:
            return {"_error": "no_access"}
        elif resp.status_code == 401:
            print("ERROR: Token expired or invalid!")
            sys.exit(1)
        else:
            return {"_error": f"http_{resp.status_code}"}
    except Exception as e:
        return {"_error": str(e)}


def get_companies():
    """Get all companies."""
    data = api_get("/v3/llm/companies")
    if isinstance(data, list):
        return [c for c in data if c["id"] not in SKIP_COMPANIES]
    return []


def get_sites(company_id):
    """Get sites for a company."""
    data = api_get(f"/v3/companies/{company_id}/sites")
    if isinstance(data, list):
        return data
    return []


def get_catalog(company_id):
    """Get data gathering catalog with availability."""
    data = api_get(
        f"/v3/data-gathering/{company_id}/catalog",
        params={"includeAvailability": "true"}
    )
    if isinstance(data, list):
        return data
    return []


def get_itf_projects(company_id):
    """Get inside-the-fence projects."""
    data = api_get(f"/v3/companies/{company_id}/inside-the-fence/projects")
    if isinstance(data, list):
        return data
    return []


def get_targets(company_id):
    """Get target tracking targets."""
    data = api_get(
        f"/v3/target-tracking/v2/{company_id}/targets",
        params={"limit": 100}
    )
    if isinstance(data, dict):
        return data.get("targets", [])
    if isinstance(data, list):
        return data
    return []


def get_scenarios(company_id):
    """Get target tracking scenarios."""
    data = api_get(
        f"/v3/target-tracking/v2/{company_id}/scenarios",
        params={"limit": 100}
    )
    if isinstance(data, dict):
        return data.get("scenarios", data.get("items", []))
    if isinstance(data, list):
        return data
    return []


def analyze_catalog(catalog):
    """Analyze catalog to extract variable info and data availability."""
    variables_with_data = []
    variables_without_data = []
    categories = set()
    min_year = None
    max_year = None

    for var in catalog:
        avail = var.get("availability")
        cat = var.get("category", "unknown")
        title = var.get("title", var.get("variableId", "?"))
        var_type = var.get("type", "raw")

        if avail and avail.get("from"):
            variables_with_data.append({
                "id": var.get("variableId"),
                "title": title,
                "category": cat,
                "type": var_type,
                "from": avail["from"],
                "to": avail.get("to", "?")
            })
            categories.add(cat)

            # Parse year from YYYY-MM format
            from_year = int(avail["from"][:4])
            to_year = int(avail.get("to", avail["from"])[:4])
            if min_year is None or from_year < min_year:
                min_year = from_year
            if max_year is None or to_year > max_year:
                max_year = to_year
        else:
            variables_without_data.append(title)

    return {
        "with_data": variables_with_data,
        "without_data": variables_without_data,
        "categories": sorted(c for c in categories if c),
        "year_range": f"{min_year}-{max_year}" if min_year else "N/A",
        "min_year": min_year,
        "max_year": max_year
    }


def analyze_projects(projects):
    """Analyze ITF projects."""
    if not projects:
        return {"count": 0, "phases": {}, "with_impact": 0, "with_start_date": 0, "metrics": set()}

    phases = {}
    with_impact = 0
    with_start_date = 0
    metrics = set()

    for p in projects:
        phase = p.get("phase", "unknown")
        phases[phase] = phases.get(phase, 0) + 1

        fields = p.get("fields", {})

        # Check for impact metrics
        impact = fields.get("impact", {})
        if impact and impact.get("value"):
            with_impact += 1
            # Try to get metric info
            metric_id = impact.get("value", {}).get("metricId") if isinstance(impact.get("value"), dict) else None
            if metric_id:
                metrics.add(metric_id)

        # Check for start date
        start = fields.get("startDate", {})
        if start and start.get("value"):
            with_start_date += 1

    return {
        "count": len(projects),
        "phases": phases,
        "with_impact": with_impact,
        "with_start_date": with_start_date,
        "metrics": metrics
    }


def compute_readiness(catalog_info, project_info, targets, sites_count):
    """
    Compute readiness score and status.
    Criteria:
    - Has data loaded in DG (variables with data > 0)
    - Has at least 2 years of data
    - Has sites configured
    - Has projects with impact defined
    - Bonus: already has targets
    """
    score = 0
    reasons = []
    blockers = []

    # Data availability (max 40 points)
    vars_with_data = len(catalog_info["with_data"])
    if vars_with_data >= 5:
        score += 40
        reasons.append(f"{vars_with_data} variables con data")
    elif vars_with_data >= 2:
        score += 25
        reasons.append(f"{vars_with_data} variables con data")
    elif vars_with_data >= 1:
        score += 10
        reasons.append(f"{vars_with_data} variable con data")
    else:
        blockers.append("Sin data en Data Gathering")

    # Year coverage (max 20 points)
    if catalog_info["min_year"] and catalog_info["max_year"]:
        year_span = catalog_info["max_year"] - catalog_info["min_year"] + 1
        if year_span >= 3:
            score += 20
            reasons.append(f"{year_span} años de data ({catalog_info['year_range']})")
        elif year_span >= 2:
            score += 15
            reasons.append(f"{year_span} años de data")
        else:
            score += 5
            reasons.append("Solo 1 año de data")
    else:
        blockers.append("Sin rango de años")

    # Sites (max 10 points)
    if sites_count > 0:
        score += 10
        reasons.append(f"{sites_count} sites")
    else:
        blockers.append("Sin sites")

    # Projects (max 20 points)
    if project_info["count"] >= 3:
        score += 15
        if project_info["with_impact"] > 0:
            score += 5
            reasons.append(f"{project_info['count']} proyectos ({project_info['with_impact']} con impacto)")
        else:
            reasons.append(f"{project_info['count']} proyectos (sin impacto definido)")
    elif project_info["count"] >= 1:
        score += 8
        reasons.append(f"{project_info['count']} proyecto(s)")
    else:
        blockers.append("Sin proyectos ITF cargados")

    # Targets bonus (max 10 points)
    if len(targets) > 0:
        score += 10
        reasons.append(f"Ya tiene {len(targets)} target(s)")

    # Determine status
    if score >= 70:
        status = "READY"
    elif score >= 40:
        status = "PARTIAL"
    else:
        status = "NOT READY"

    return {
        "score": score,
        "status": status,
        "reasons": reasons,
        "blockers": blockers
    }


def main():
    if not TOKEN:
        print("Usage: python3 audit_tt_readiness.py <bearer_token>")
        print("  Pass the token WITHOUT 'Bearer ' prefix")
        sys.exit(1)

    print("=" * 80)
    print("TARGET TRACKING READINESS AUDIT")
    print(f"Date: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    print("=" * 80)

    companies = get_companies()
    print(f"\nFound {len(companies)} companies (excluding demos/sandboxes)\n")

    results = []

    for i, company in enumerate(companies):
        cid = company["id"]
        cname = company["name"]
        print(f"[{i+1}/{len(companies)}] Auditing {cname} ({cid})...", end=" ", flush=True)

        # Fetch all data
        sites = get_sites(cid)
        catalog = get_catalog(cid)
        projects = get_itf_projects(cid)
        targets = get_targets(cid)
        scenarios = get_scenarios(cid)

        # Analyze
        catalog_info = analyze_catalog(catalog)
        project_info = analyze_projects(projects)
        readiness = compute_readiness(catalog_info, project_info, targets, len(sites))

        result = {
            "company_id": cid,
            "company_name": cname.strip(),
            "sites_count": len(sites),
            "sites_names": [s.get("name", s.get("id", "?")) for s in sites],
            "variables_with_data": len(catalog_info["with_data"]),
            "variables_without_data": len(catalog_info["without_data"]),
            "categories": ", ".join(catalog_info["categories"]),
            "year_range": catalog_info["year_range"],
            "min_year": catalog_info["min_year"],
            "max_year": catalog_info["max_year"],
            "variable_details": catalog_info["with_data"],
            "projects_count": project_info["count"],
            "projects_phases": project_info["phases"],
            "projects_with_impact": project_info["with_impact"],
            "projects_with_start_date": project_info["with_start_date"],
            "targets_count": len(targets),
            "target_names": [t.get("name", t.get("metricId", "?")) for t in targets],
            "scenarios_count": len(scenarios) if isinstance(scenarios, list) else 0,
            "readiness_score": readiness["score"],
            "readiness_status": readiness["status"],
            "readiness_reasons": readiness["reasons"],
            "readiness_blockers": readiness["blockers"],
        }

        results.append(result)
        status_icon = {"READY": "✅", "PARTIAL": "🟡", "NOT READY": "❌"}.get(readiness["status"], "?")
        print(f"{status_icon} {readiness['status']} (score: {readiness['score']})")

        # Small delay to avoid rate limiting
        time.sleep(0.3)

    # Sort by readiness score descending
    results.sort(key=lambda r: r["readiness_score"], reverse=True)

    # Print summary
    print("\n" + "=" * 80)
    print("RESUMEN DE READINESS")
    print("=" * 80)

    ready = [r for r in results if r["readiness_status"] == "READY"]
    partial = [r for r in results if r["readiness_status"] == "PARTIAL"]
    not_ready = [r for r in results if r["readiness_status"] == "NOT READY"]

    print(f"\n✅ READY ({len(ready)}):")
    for r in ready:
        print(f"   {r['company_name']:<35} Score: {r['readiness_score']:>3} | "
              f"Sites: {r['sites_count']} | Vars: {r['variables_with_data']} | "
              f"Years: {r['year_range']} | Projects: {r['projects_count']} | "
              f"Targets: {r['targets_count']}")

    print(f"\n🟡 PARTIAL ({len(partial)}):")
    for r in partial:
        blockers = " | Blockers: " + "; ".join(r["readiness_blockers"]) if r["readiness_blockers"] else ""
        print(f"   {r['company_name']:<35} Score: {r['readiness_score']:>3} | "
              f"Sites: {r['sites_count']} | Vars: {r['variables_with_data']} | "
              f"Years: {r['year_range']} | Projects: {r['projects_count']}{blockers}")

    print(f"\n❌ NOT READY ({len(not_ready)}):")
    for r in not_ready:
        blockers = "; ".join(r["readiness_blockers"])
        print(f"   {r['company_name']:<35} Score: {r['readiness_score']:>3} | {blockers}")

    # Export to Excel
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = openpyxl.Workbook()

        # --- Sheet 1: Summary ---
        ws = wb.active
        ws.title = "TT Readiness Summary"

        headers = [
            "Company", "Status", "Score", "Sites", "Variables w/ Data",
            "Categories", "Year Range", "Projects ITF", "Projects w/ Impact",
            "Targets Existentes", "Scenarios", "Blockers", "Next Steps"
        ]

        # Header styling
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        ready_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        partial_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        not_ready_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        for row_idx, r in enumerate(results, 2):
            # Determine next steps
            next_steps = []
            if r["readiness_status"] == "READY":
                if r["targets_count"] == 0:
                    next_steps.append("Crear targets en plataforma")
                if r["projects_count"] == 0:
                    next_steps.append("Cargar proyectos del cliente")
                next_steps.append("Agendar demo TT con cliente + Ivo")
            elif r["readiness_status"] == "PARTIAL":
                if r["variables_with_data"] == 0:
                    next_steps.append("Subir data con Tommy Gómez")
                if r["projects_count"] == 0:
                    next_steps.append("Cargar proyectos (reales o ejemplo)")
                next_steps.append("Evaluar viabilidad de demo parcial")
            else:
                next_steps.append("Requiere data gathering primero")

            values = [
                r["company_name"],
                r["readiness_status"],
                r["readiness_score"],
                r["sites_count"],
                r["variables_with_data"],
                r["categories"],
                r["year_range"],
                r["projects_count"],
                r["projects_with_impact"],
                r["targets_count"],
                r["scenarios_count"],
                "; ".join(r["readiness_blockers"]),
                "; ".join(next_steps)
            ]

            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.border = thin_border
                cell.alignment = Alignment(wrap_text=True)

            # Color-code status
            status_cell = ws.cell(row=row_idx, column=2)
            if r["readiness_status"] == "READY":
                status_cell.fill = ready_fill
            elif r["readiness_status"] == "PARTIAL":
                status_cell.fill = partial_fill
            else:
                status_cell.fill = not_ready_fill

        # Auto-width
        for col in ws.columns:
            max_len = 0
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

        # Freeze header
        ws.freeze_panes = "A2"

        # --- Sheet 2: Variable Details ---
        ws2 = wb.create_sheet("Variable Details")
        var_headers = ["Company", "Variable ID", "Variable Title", "Category", "Type", "Data From", "Data To"]

        for col, header in enumerate(var_headers, 1):
            cell = ws2.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border

        row_idx = 2
        for r in results:
            for v in r.get("variable_details", []):
                vals = [
                    r["company_name"], v["id"], v["title"],
                    v["category"], v["type"], v["from"], v["to"]
                ]
                for col, val in enumerate(vals, 1):
                    cell = ws2.cell(row=row_idx, column=col, value=val)
                    cell.border = thin_border
                row_idx += 1

        ws2.freeze_panes = "A2"

        # --- Sheet 3: Action Plan for SSPs ---
        ws3 = wb.create_sheet("SSP Action Plan")
        action_headers = [
            "Company", "SSP Owner", "Status", "Score",
            "Paso 1: Validar Data", "Paso 2: Cargar Proyectos",
            "Paso 3: Crear Targets", "Paso 4: Demo con Cliente",
            "Fecha Límite Demo", "Notas"
        ]

        for col, header in enumerate(action_headers, 1):
            cell = ws3.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border

        row_idx = 2
        for r in [x for x in results if x["readiness_status"] in ("READY", "PARTIAL")]:
            data_ok = "✅ Data cargada" if r["variables_with_data"] >= 2 else "⚠️ Falta data"
            proj_ok = f"✅ {r['projects_count']} proyectos" if r["projects_count"] >= 1 else "❌ Cargar proyectos"
            tgt_ok = f"✅ {r['targets_count']} targets" if r["targets_count"] >= 1 else "❌ Crear targets"
            demo = "Agendar con Ivo" if r["readiness_status"] == "READY" else "Evaluar viabilidad"

            vals = [
                r["company_name"], "(asignar)", r["readiness_status"], r["readiness_score"],
                data_ok, proj_ok, tgt_ok, demo, "", ""
            ]
            for col, val in enumerate(vals, 1):
                cell = ws3.cell(row=row_idx, column=col, value=val)
                cell.border = thin_border
                cell.alignment = Alignment(wrap_text=True)
            row_idx += 1

        ws3.freeze_panes = "A2"
        for col in ws3.columns:
            ws3.column_dimensions[col[0].column_letter].width = 22

        # Save
        output_dir = Path(__file__).parent / "outputs"
        output_dir.mkdir(exist_ok=True)
        output_path = output_dir / f"TT_Readiness_Audit_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        wb.save(str(output_path))
        print(f"\n📊 Excel exported to: {output_path}")

    except ImportError:
        print("\n⚠️  openpyxl not installed. Run: pip3 install openpyxl")
        # Fallback: save JSON
        output_path = Path(__file__).parent / "outputs" / f"TT_Readiness_Audit_{datetime.now().strftime('%Y%m%d_%H%M')}.json"
        output_path.parent.mkdir(exist_ok=True)
        with open(output_path, "w") as f:
            json.dump(results, f, indent=2, default=str)
        print(f"   JSON exported to: {output_path}")

    # Also save raw JSON for programmatic use
    json_path = Path(__file__).parent / "outputs" / f"TT_Readiness_Raw_{datetime.now().strftime('%Y%m%d_%H%M')}.json"
    with open(json_path, "w") as f:
        json.dump(results, f, indent=2, default=str)
    print(f"📄 Raw JSON: {json_path}")


if __name__ == "__main__":
    main()
